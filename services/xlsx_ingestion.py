"""
services/xlsx_ingestion.py
==========================
Motor de Ingestão de Planilhas Excel Locais (PEDIDOS_SAMBA_.xlsx) - Nível Enterprise.
Arquitetura tolerante a falhas humanas com Extração Heurística via Regex Avançada.
Prepara o terreno para Agentes de IA complementarem os dados estruturados.
"""

import sys
import logging
import re
import json
import hashlib
import openpyxl
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Any, Optional

# Garantir path absoluto
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from models.database import get_session, create_tables, Deal, StrategicData

logging.basicConfig(level=logging.INFO, format='%(levelname)s: [%(name)s] %(message)s')
logger = logging.getLogger("ExcelIngestionEngine")

# ============================================================================
# 1. DICIONÁRIOS DE CONHECIMENTO CORPORATIVO (KNOWLEDGE BASE)
# ============================================================================

XLSX_PATH = ROOT / "data" / "sheets" / "PEDIDOS_SAMBA_.xlsx"

# Arrays de Contexto de Negócio
INCOTERMS = ["FOB", "CIF", "CFR", "EXW", "DAP", "DDP", "FAS", "CPT", "CIP", "ASWP"]
PAYMENT_TERMS = ["DLC", "SBLC", "TT", "LC", "CAD", "BG", "MT103", "ESCROW"]
DESTINATIONS = [
    "China", "Vietna", "Vietnam", "Haiphong", "Bangladesh", "Israel", 
    "Egito", "Egypt", "Rotterdam", "Dakar", "ASWP", "Rep. Dominicana", 
    "Mauricio", "Brasil", "Brazil", "Dubai", "UAE", "India", "USA"
]

# Grupos para Inferência de Direção (Bid vs Ask)
GRUPOS_FORNECEDORES = ["rokane", "conex", "fabricio", "gerson", "ze vasconcelos", "primex", "barreto", "mel", "usina"]
GRUPOS_COMPRADORES = ["bahov", "gwi", "dionathan", "maxin", "bicca", "dannyel", "bruno", "vilson", "ariel", "kent foods"]

# Atribuição de Sócios (Rule-based Routing)
ASSIGNEE_RULES = [
    (["soja", "milho", "trigo", "arroz", "feijão", "feijao", "sorgo", "farelo", "corn", "soy", "wheat"], "Leonardo"),
    (["açúcar", "acucar", "ic45", "icumsa", "etanol", "algodão", "algodao", "sugar", "vhp"], "Nivio"),
    (["café", "cafe", "cacau", "frango", "chicken", "boi", "beef", "porco", "pork", "oleo", "óleo", "oil", "tallow", "paw"], "Marcelo")
]

# Mapeamento do Status Humano -> Estágio do Pipeline Kanban
STATUS_PIPELINE_MAP = {
    "pendente vendedor": "Negociação",
    "pendente comprador": "Negociação",
    "emissão icpo": "Proposta Enviada",
    "emissão de icpo": "Proposta Enviada",
    "emissao icpo": "Proposta Enviada",
    "fco": "Proposta Enviada",
    "spa": "Contrato Assinado",
    "sblc": "Em Due Diligence",
    "dlc": "Em Due Diligence",
    "finder": "Qualificação",
    "mandato": "Qualificação",
    "parada": "Lead Capturado",
    "declinado": "Declinado",
    "cancelado": "Declinado",
    "fechado": "Fechado"
}

ARCHIVED_STATUSES = ["declinado", "cancelado", "fechado", "parada"]
IGNORE_SHEETS = ["Para Vender", "Ariel", "Agenda", "ACUCAR", "OLEO", "proteinas + ovo", "Linkedin"]


# ============================================================================
# 2. MOTOR DE SANITIZAÇÃO DE DADOS (DATA SANITIZER)
# ============================================================================

class DataSanitizer:
    """Classe utilitária para limpar a sujeira humana da planilha."""

    @staticmethod
    def cell_to_str(val: Any) -> str:
        if val is None: return ""
        return str(val).strip()

    @staticmethod
    def parse_date(val: Any) -> datetime:
        if isinstance(val, datetime): return val
        try:
            return datetime.fromisoformat(str(val)) if val else datetime.utcnow()
        except Exception:
            return datetime.utcnow()

    @staticmethod
    def parse_br_number(val: str) -> Optional[float]:
        """Lida com as loucuras de formatação numérica do Excel/Humano."""
        if not val: return None
        s = str(val).strip().lower()
        s = re.sub(r"[^\d.,-]", "", s) # Mantém apenas números, ponto, vírgula e sinal
        
        if not s: return None

        # Exemplo: 3.250.000,00 -> 3250000.00
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."): # Vírgula é decimal (BR)
                s = s.replace(".", "").replace(",", ".")
            else: # Ponto é decimal (US)
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".") # 975,50 -> 975.50
        elif "." in s and s.count(".") == 1:
            parts = s.split(".")
            # Se a parte "decimal" tem 3 dígitos cravados, é quase certeza que o humano usou ponto como milhar (Ex: 25.000)
            if len(parts[1]) == 3:
                s = s.replace(".", "")
            # Caso contrário, mantém como decimal (Ex: 3.5)

        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def normalize_commodity(raw: str) -> str:
        """Taxonomia estrita para permitir relatórios consolidados futuros."""
        if not raw: return "Indefinida"
        r_norm = DataSanitizer._remove_accents(raw.strip().lower())
        
        TAXONOMY_MAP = {
            "acucar ic45": ["ic45", "icumsa 45", "branco refinado"],
            "Açúcar VHP": ["icumsa vhp", "vhp", "bruto", "mascavo"],
            "Açúcar": ["acucar", "sugar"],
            "Soja": ["soja", "soybean", "soy"],
            "Farelo de Soja": ["farelo", "meal"],
            "Milho Amarelo": ["milho amarelo", "yellow corn", "gmo", "non gmo"],
            "Milho Branco": ["milho branco", "white corn"],
            "Milho": ["milho", "corn"],
            "Arroz": ["arroz", "rice"],
            "Café": ["cafe", "coffee"],
            "Cacau": ["cacau", "cocoa"],
            "Algodão": ["algodao", "cotton"],
            "Chicken Paw": ["chicken", "paw", "pe de frango"],
            "Frango": ["frango", "poultry"],
            "Carne Suína": ["porco", "pork", "suino", "feet"],
            "Carne Bovina": ["boi", "beef", "tallow", "bovino", "acem"],
            "Etanol": ["etanol", "ethanol"],
            "Óleo de Soja": ["soja refin", "soybean refin", "oleo soja"],
            "Óleo de Girassol": ["girassol", "sunflower"],
            "Óleo de Palma": ["palma", "palm"],
            "Óleo Vegetal": ["oleo", "oil"],
            "Indefinida": ["conexao", "reuniao", "diversos"]
        }

        for canonical, keywords in TAXONOMY_MAP.items():
            if any(kw in r_norm for kw in keywords):
                return canonical
        
        return raw.strip().title()[:60]

    @staticmethod
    def _remove_accents(text: str) -> str:
        accents = {'ã':'a', 'á':'a', 'à':'a', 'â':'a', 'é':'e', 'ê':'e', 'í':'i', 'ó':'o', 'õ':'o', 'ô':'o', 'ú':'u', 'ç':'c'}
        for k, v in accents.items():
            text = text.replace(k, v)
        return text


# ============================================================================
# 3. HEURÍSTICAS DE INTELIGÊNCIA ARTIFICIAL E REGEX (AI PARSER)
# ============================================================================

class ContextHeuristics:
    """
    Simula o raciocínio humano para entender textos caóticos e inferir 
    fatos não declarados explicitamente na planilha.
    """

    @staticmethod
    def infer_direction(oferta_pedido: str, grupo: str, fornecedor: str, comprador: str) -> str:
        """Determina se a Samba Export está na ponta de Compra (BID) ou Venda (ASK)."""
        op = DataSanitizer.cell_to_str(oferta_pedido).lower()
        if "pedido" in op: return "BID"
        if "oferta" in op: return "ASK"
        
        grp = DataSanitizer.cell_to_str(grupo).lower()
        if any(f in grp for f in GRUPOS_FORNECEDORES): return "ASK"
        if any(c in grp for c in GRUPOS_COMPRADORES): return "BID"

        # Se houver fornecedor mapeado e não tiver comprador, provavelmente é uma oferta chegando
        if fornecedor and not comprador: return "ASK"
        if comprador and not fornecedor: return "BID"
        
        return "UNKNOWN"

    @staticmethod
    def infer_stage(status_raw: str) -> str:
        key = DataSanitizer.cell_to_str(status_raw).lower()
        for k, v in STATUS_PIPELINE_MAP.items():
            if k in key: return v
        return "Lead Capturado"

    @staticmethod
    def extract_trade_parameters(text_j: str, text_l: str) -> Dict[str, Any]:
        """
        O coração do sistema. Varre anotações livres em busca de parâmetros contratuais.
        Retorna: price, volume, currency, incoterm, origin, destination, payment_terms
        """
        full_text = f"{DataSanitizer.cell_to_str(text_j)} {DataSanitizer.cell_to_str(text_l)}".upper()
        
        params = {
            "price": None, "volume": None, "currency": "USD",
            "incoterm": None, "origin": None, "destination": None, "payment": None
        }

        if not full_text.strip():
            return params

        # 1. Incoterms & Payment Terms
        m_inc = re.search(r"\b(" + "|".join(INCOTERMS) + r")\b", full_text)
        if m_inc: params["incoterm"] = m_inc.group(1)

        m_pay = re.search(r"\b(" + "|".join(PAYMENT_TERMS) + r")\b", full_text)
        if m_pay: params["payment"] = m_pay.group(1)

        # 2. Volume (Identifica "MT", "Tons", Sacas e Milhares)
        # Tenta pegar milhares: "300 mil mt", "50k tons"
        m_vol = re.search(r"(\d[\d.,]*)\s*(MIL|K|MM)?\s*(MT|TONS?|T\b|SACAS?|SC)", full_text)
        if m_vol:
            v_val = DataSanitizer.parse_br_number(m_vol.group(1))
            if v_val:
                multiplier = 1000 if m_vol.group(2) in ["MIL", "K"] else 1000000 if m_vol.group(2) == "MM" else 1
                params["volume"] = v_val * multiplier
                # Nota: Deixamos na base como MT, independentemente de ser saca. O frontend lida com a saca.

        # Fallback volume isolado: "25,000 MT"
        if not params["volume"]:
            m_vol_dir = re.search(r"([\d]{1,3}(?:,\d{3})+|[\d]{1,3}(?:\.\d{3})+)\s*(?:MT|TONS?|T\b)", full_text)
            if m_vol_dir:
                params["volume"] = DataSanitizer.parse_br_number(m_vol_dir.group(1))

        # 3. Price & Currency
        # Match: "355 USD", "USD 355", "$355", "3690 usd"
        m_usd = re.search(r"(\d[\d.,]*)\s*(USD|US\$|U\$)|(?:USD|US\$)\s*(\d[\d.,]*)", full_text)
        if m_usd:
            p_val = DataSanitizer.parse_br_number(m_usd.group(1) or m_usd.group(3))
            if p_val and 10 <= p_val <= 100000:
                params["price"], params["currency"] = p_val, "USD"
        else:
            # Fallback para Reais: "R$ 130", "130 BRL"
            m_brl = re.search(r"(\d[\d.,]*)\s*(BRL|REAIS)|(?:BRL|R\$)\s*(\d[\d.,]*)", full_text)
            if m_brl:
                p_val = DataSanitizer.parse_br_number(m_brl.group(1) or m_brl.group(3))
                if p_val and 10 <= p_val <= 100000:
                    params["price"], params["currency"] = p_val, "BRL"

        # 4. Location Context (Origem e Destino)
        if params["incoterm"] and params["incoterm"] != "ASWP":
            # Geralmente o destino segue o CIF/CFR (Ex: "CIF China")
            m_dest = re.search(rf"{params['incoterm']}\s+([A-ZÀ-ÖØ-öø-ÿ]+)", full_text)
            if m_dest: params["destination"] = m_dest.group(1).title()

        m_orig = re.search(r"ORIGEM\s+([A-ZÀ-ÖØ-öø-ÿ]+)", full_text)
        if m_orig: params["origin"] = m_orig.group(1).title()

        # Fallback de destino por Keyword
        if not params["destination"]:
            for d in DESTINATIONS:
                if re.search(r"\b" + re.escape(d.upper()) + r"\b", full_text):
                    params["destination"] = d
                    break

        return params


# ============================================================================
# 4. FÁBRICA DE DEALS (PIPELINE BUILDER)
# ============================================================================

class DealFactory:
    """Responsável por orquestrar a montagem do objeto e criar Hashes de Integridade."""

    @staticmethod
    def generate_hash(date_obj: datetime, commodity: str, group: str, vol: float) -> str:
        """Cria uma assinatura digital da linha para validação de duplicidade avançada."""
        d_str = date_obj.strftime("%Y%m%d") if date_obj else "00000000"
        c_str = DataSanitizer.cell_to_str(commodity).lower()[:10]
        g_str = DataSanitizer.cell_to_str(group).lower()[:10]
        v_str = str(int(vol)) if vol else "0"
        raw_string = f"{d_str}_{c_str}_{g_str}_{v_str}"
        return hashlib.md5(raw_string.encode()).hexdigest()[:10]

    @classmethod
    def create_from_row(cls, sheet_name: str, row_num: int, raw_row: tuple, is_2025: bool=False, force_declinado: bool=False) -> Optional[Dict]:
        """
        Padroniza a leitura das colunas:
        A=JOB(0), B=DATA(1), C=OFERTA/PEDIDO(2), D=GRUPO(3), E=SOLICIT(4), F=STATUS(5),
        G=PRODUTO(6), H=COMPRADOR(7), I=FORNECEDOR(8), J=VIZ_RAPIDA(9), K=DOCS(10), L=SPEC(11), N=FUP(13)
        """
        r = list(raw_row) + [""] * 20 # Previne IndexError em linhas curtas

        job, data_raw, op, grupo, solicit, status = r[0], r[1], r[2], r[3], r[4], r[5]
        produto, comprador, fornec = r[6], r[7], r[8]
        viz, docs, spec, fup = r[9], r[10], r[11], r[13]

        job_str = DataSanitizer.cell_to_str(job)
        prod_str = DataSanitizer.cell_to_str(produto)
        
        # Ignora lixos e subtotais
        if not prod_str and not DataSanitizer.cell_to_str(grupo) and not job_str: return None
        if prod_str.upper() in ("TOTAL", "TOTAIS"): return None

        # Datas e Categorias
        created_at = DataSanitizer.parse_date(data_raw)
        direcao = ContextHeuristics.infer_direction(op, grupo, fornec, comprador)
        stage = ContextHeuristics.infer_stage(status)
        deal_status = "declinado" if force_declinado or DataSanitizer.cell_to_str(status).lower() in ARCHIVED_STATUSES else "ativo"
        commodity_norm = DataSanitizer.normalize_commodity(prod_str)

        # Extração de Inteligência (A Mágica)
        params = ContextHeuristics.extract_trade_parameters(viz, spec)
        
        # Consolidação de Identificadores e Notas
        # Salva o RAW JSON nas notes para agentes LLMs futuros poderem analisar falhas humanas!
        raw_context = {
            "viz_rapida": DataSanitizer.cell_to_str(viz),
            "especificacao": DataSanitizer.cell_to_str(spec),
            "follow_up": DataSanitizer.cell_to_str(fup),
            "docs": DataSanitizer.cell_to_str(docs),
            "pagamento_inferido": params["payment"]
        }
        if is_2025: raw_context["historico"] = True

        # Cria ID único se o humano não tiver preenchido a coluna JOB
        deal_hash = cls.generate_hash(created_at, commodity_norm, grupo, params["volume"])
        name = job_str if job_str else f"XLSX_{sheet_name[:3].upper()}_{deal_hash}_{row_num:03d}"

        # Atribuição do sócio responsável baseada em regras
        assignee = next((s for kws, s in ASSIGNEE_RULES if any(kw in commodity_norm.lower() for kw in kws)), "Leonardo")

        return dict(
            name=name,
            commodity=commodity_norm,
            direcao=direcao,
            stage=stage,
            status=deal_status,
            price=params["price"],
            volume=params["volume"],
            volume_unit="MT" if params["volume"] else None,
            currency=params["currency"],
            incoterm=params["incoterm"],
            origin=params["origin"] or DataSanitizer.cell_to_str(fornec) or None,
            destination=params["destination"] or DataSanitizer.cell_to_str(comprador) or None,
            source_group=DataSanitizer.cell_to_str(grupo) or sheet_name,
            source_sender=DataSanitizer.cell_to_str(solicit) or DataSanitizer.cell_to_str(comprador) or DataSanitizer.cell_to_str(fornec),
            assignee=assignee,
            notes=f"[XLSX][{sheet_name}] RAW_DATA:\n{json.dumps(raw_context, ensure_ascii=False, indent=2)}",
            created_at=created_at,
            risk_score=50,
            alerta_grupo_interno=deal_hash # Usamos esse campo temporariamente para guardar o Hash Antiduplicata
        )


# ============================================================================
# 5. ESTRATÉGIAS DE PARSING POR TIPO DE ABA
# ============================================================================

class SheetParsers:

    @staticmethod
    def parse_standard_pipeline(ws, sheet_name: str, force_declinado: bool = False, is_2025: bool = False) -> List[Dict]:
        """Engine para Abas de Pipeline convencionais (dados a partir da linha 5)."""
        deals = []
        for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            deal_obj = DealFactory.create_from_row(sheet_name, i, row, is_2025, force_declinado)
            if deal_obj: deals.append(deal_obj)
        return deals

    @staticmethod
    def parse_gwi_specific(ws) -> List[Dict]:
        """Engine Customizado para a estrutura 'oleo GWI'."""
        deals = []
        data_gwi = datetime(2026, 3, 31)
        for i, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
            r = list(row) + [""] * 20
            produto, trial_vol, embalagem, destino, preco_raw, status_n = r[0], r[1], r[3], r[4], r[5], r[13]
            
            p_str = DataSanitizer.cell_to_str(produto)
            if not p_str or p_str.upper() in ("TOTAL", "PRODUTO", ""): continue
            
            vol = DataSanitizer.parse_br_number(str(trial_vol)) if trial_vol else None
            
            price, currency = None, "USD"
            m = re.search(r"USD\s*([\d][0-9.,]*)", DataSanitizer.cell_to_str(preco_raw), re.IGNORECASE)
            if m:
                p_val = DataSanitizer.parse_br_number(m.group(1))
                if p_val and 10 < p_val < 1000000: price = p_val
                
            comm = DataSanitizer.normalize_commodity(p_str)
            dest = DataSanitizer.cell_to_str(destino)
            name = f"GWIOIL_{re.sub(r'[^A-Z0-9]', '', p_str.upper())[:15]}_{i:03d}"
            stage = "Proposta Enviada" if "SPA" in DataSanitizer.cell_to_str(status_n).upper() else "Qualificação"
            
            raw_context = {"embalagem": DataSanitizer.cell_to_str(embalagem), "status_parceiro": DataSanitizer.cell_to_str(status_n)}

            # CORREÇÃO 2: Atribuição correta do assignee usando a list comprehension inline para o contexto GWI
            assignee = next((s for kws, s in ASSIGNEE_RULES if any(kw in comm.lower() for kw in kws)), "Marcelo")

            deals.append(dict(
                name=name, commodity=comm, direcao="BID", stage=stage, status="ativo",
                price=price, volume=vol, volume_unit="MT", currency=currency, incoterm="CIF",
                destination=dest, source_group="GWI", source_sender="GDK - Kent Foods",
                assignee=assignee, 
                notes=f"[XLSX][GWI] RAW_DATA:\n{json.dumps(raw_context, ensure_ascii=False)}",
                created_at=data_gwi, risk_score=50, alerta_grupo_interno="GWI_SPECIAL"
            ))
        return deals

    @staticmethod
    def parse_strategic_catalog(ws, sheet_name: str) -> List[StrategicData]:
        """Engine Consultiva para transformar Grades em JSON Vetorizável."""
        out = []
        hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [DataSanitizer.cell_to_str(h) for h in hdr_row]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [DataSanitizer.cell_to_str(v) for v in row]
            if not any(vals): continue
            
            record = {}
            for i, val in enumerate(vals):
                if i < len(headers) and headers[i] and val:
                    record[headers[i]] = val
                    
            if record:
                out.append(StrategicData(sheet_name=sheet_name, raw_json=json.dumps(record, ensure_ascii=False)))
        return out


# ============================================================================
# 6. ORQUESTRAÇÃO DE BANCO E EXECUÇÃO PRINCIPAL
# ============================================================================

def _upsert_deals(session, deals: List[Dict], existing_names: set, existing_hashes: set) -> Dict[str, int]:
    """Insere preservando a idempotência."""
    stats = {"inserted": 0, "skipped_name": 0, "skipped_hash": 0, "errors": 0}
    
    for d in deals:
        nm = d.get("name")
        deal_hash = d.pop("alerta_grupo_interno", None) # Removemos o hash temporário do dicionário

        if nm in existing_names:
            stats["skipped_name"] += 1
            continue
        
        # Filtro de Hash (Impede o mesmo deal de ser criado com IDs genéricos diferentes)
        if deal_hash and deal_hash in existing_hashes and deal_hash not in ("GWI_SPECIAL", "00000000_indefinida__0"):
            stats["skipped_hash"] += 1
            continue

        try:
            deal = Deal(**d)
            session.add(deal)
            existing_names.add(nm)
            if deal_hash: existing_hashes.add(deal_hash)
            stats["inserted"] += 1
        except Exception as e:
            logger.error(f"Erro SQL ao inserir Deal [{nm}]: {e}")
            stats["errors"] += 1

    try:
        session.commit()
    except Exception as e:
        session.rollback()
        logger.error(f"Falha de Commit no Lote: {e}")
        stats["errors"] += stats["inserted"]
        stats["inserted"] = 0
        
    return stats


def ingest_xlsx():
    """O Cérebro Mestre da Ingestão."""
    logger.info("="*60)
    logger.info(f"🚀 INICIANDO INGESTÃO ENTERPRISE DO ARQUIVO EXCEL")
    logger.info(f"📄 Path: {XLSX_PATH}")
    logger.info("="*60)
    
    create_tables()
    session = get_session()

    # 1. Limpeza Cirúrgica (apenas os dados gerados por este script)
    deleted_deals = session.query(Deal).filter(Deal.notes.like("%[XLSX]%")).delete(synchronize_session=False)
    session.query(StrategicData).delete(synchronize_session=False)
    session.commit()
    logger.info(f"🧹 PURGE: {deleted_deals} Deals antigos de planilhas removidos. Catálogos estratégicos zerados.")

    # 2. Carregamento de Cache para Deduplicação O(1)
    existing_names = set(r[0] for r in session.query(Deal.name).all() if r[0])
    # Como não temos um campo 'hash' nativo, usamos o cache de nomes para o básico, 
    # e a validação de Hash ocorrerá durante o loop da sessão atual.
    existing_hashes = set() 

    # 3. Leitura Segura do Arquivo
    try:
        wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True, data_only=True)
    except FileNotFoundError:
        logger.error(f"CRÍTICO: Arquivo não localizado no sistema de arquivos.")
        session.close()
        return

    sheets = wb.sheetnames
    logger.info(f"🔍 Abas Mapeadas ({len(sheets)}): {sheets}")

    global_stats = {"pipeline_deals": 0, "strategic_records": 0, "skipped": 0}

    # 4. Roteamento por Aba (Switch/Case Lógico)
    for sheet_name in sheets:
        # CORREÇÃO 1: Utilização da constante IGNORE_SHEETS corrigida
        if sheet_name in IGNORE_SHEETS:
            logger.info(f"⏭️  Ignorando (Regra de Negócio): {sheet_name}")
            continue

        logger.info(f"⚙️  Analisando Aba: [{sheet_name}]")
        ws = wb[sheet_name]

        # --- Fluxo de Pipeline (Geram Oportunidades Financeiras) ---
        deals_batch = []
        if sheet_name in ["todos andamento", "Andamento Vietnã"]:
            deals_batch = SheetParsers.parse_standard_pipeline(ws, sheet_name)
        elif sheet_name == "Declinados":
            deals_batch = SheetParsers.parse_standard_pipeline(ws, sheet_name, force_declinado=True)
        elif sheet_name == "2025":
            deals_batch = SheetParsers.parse_standard_pipeline(ws, sheet_name, is_2025=True)
        elif sheet_name == "oleo GWI":
            deals_batch = SheetParsers.parse_gwi_specific(ws)

        if deals_batch:
            res = _upsert_deals(session, deals_batch, existing_names, existing_hashes)
            global_stats["pipeline_deals"] += res["inserted"]
            global_stats["skipped"] += (res["skipped_name"] + res["skipped_hash"])
            logger.info(f"   ↳ Deals Inseridos: {res['inserted']} | Skips (Duplicados/Vazios): {res['skipped_name'] + res['skipped_hash']}")

        # --- Fluxo Consultivo (Geram Base de Conhecimento RAG/LLM) ---
        elif sheet_name in ["rokane", "valores comuns ", "FORNECEDORES"]:
            records = SheetParsers.parse_strategic_catalog(ws, sheet_name)
            for r in records: session.add(r)
            global_stats["strategic_records"] += len(records)
            logger.info(f"   ↳ Registros Estratégicos Catalogados: {len(records)}")

    # 5. Finalização e Relatório de Integridade
    session.commit()
    
    bid = session.query(Deal).filter(Deal.direcao=="BID").count()
    ask = session.query(Deal).filter(Deal.direcao=="ASK").count()
    unk = session.query(Deal).filter(Deal.direcao=="UNKNOWN").count()

    logger.info("==================================================")
    logger.info("✅ INGESTÃO CONCLUÍDA COM SUCESSO ABSOLUTO")
    logger.info(f"📈 Pipeline: {global_stats['pipeline_deals']} novos Deals.")
    logger.info(f"📚 Catálogo: {global_stats['strategic_records']} linhas de inteligência.")
    logger.info(f"🛡️ Bloqueios Anti-Duplicidade: {global_stats['skipped']} eventos.")
    logger.info("---")
    logger.info(f"⚖️ BALANCEAMENTO DO BANCO DE DADOS:")
    logger.info(f"   BID (Compras): {bid}")
    logger.info(f"   ASK (Vendas) : {ask}")
    logger.info(f"   UNKNOWN      : {unk}")
    logger.info("==================================================")
    
    session.close()

if __name__ == "__main__":
    ingest_xlsx()